from openpyxl import load_workbook
import os
import shutil
import re
from datetime import datetime

from apps.core.models import ModelArtic

from django.db import transaction

#RUTE_SIAAC ="./siaac-files/"
RUTE_SIAAC_FILES = "./siaac_files/"


ABC = ['', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y', 'Z']
# si el porcentaje es positivo aumenta sino hace un decuento
# retorna el numero con el porsentaje aplicado
def percent_apli(num, percent):
    aux = abs(percent / 100)
    apli_p = num * aux

    if percent > 0:
        result = num + apli_p
    elif percent < 0:
        result = num - apli_p
    else:
        result = 0

    return round(result,2)


def buscarPrecio(cod, lista_num):
    #BUSCA POR CODIGO EL PRECIO DEL ARTICULO en siaac
    long_precio = 12
    if lista_num == 1:
        inicioPrecio = 76
    if lista_num == 5:
        inicioPrecio = 124

    finprecio = inicioPrecio + long_precio
    patron = '^' + cod.upper().strip() + ' '
    file = open(RUTE_SIAAC_FILES + "articDB.txt")
    i=0
    
    
    for linea in file:
       
        result = re.findall(patron, linea)

       
        if result:
            precio = linea[inicioPrecio:finprecio]
           
            return precio.strip(" ")

    print(f"\n\n*********************************************\n ERROR: codigo {cod} no encontrado reviselo\n*********************************************\n")    
    return -1


# dado una ruta de un archivo xlsx devuelve un dic con todos los codgos descripcion, precios, y cleda
def get_artcis_from_xlsx(rute_xlsx):
    
    try:
        wb = load_workbook(rute_xlsx)
    except FileNotFoundError:
        print(f'{rute_xlsx} not found')
        return None
    sheet = wb['Hoja1']

    maxRow = sheet.max_row
    
    code_pattern = "^[A-Z]+-"

    list_codes = {}
    col = 1
    row = 1
    while row < maxRow:
        cell=str(sheet.cell(row=row,column=col).value).upper()
        cell_title = cell
        if cell_title == "COD" or cell_title == "COD.":
            row += 1
            cell=str(sheet.cell(row=row,column=col).value).upper()
            
            # valido q sea un codigo. el codigo se compone de letras mayusculas seguidas por un guion (ej: TIR-250)
            result = re.findall(code_pattern, cell)
            
            while result:

                
                price = sheet.cell(row=row,column=col+3).value
                if isinstance(price, str):
                    try:
                        price = float(price.strip())
                    except ValueError:
                        price = None
                list_codes[cell] = {
                    'description': str(sheet.cell(row=row,column=col+1).value).upper(),
                    'price': price,
                    'row': row,
                    'col': col
                }
                

                row += 1
                cell=str(sheet.cell(row=row,column=col).value).strip().upper()
                if cell:                    
                    result = re.findall(code_pattern, cell.upper())
                else:
                    result = None
        if cell == 'COD' or cell == 'COD.':
            pass
        else:
            row += 1

    return list_codes

def copy_file(origin, folder):
    # Obtener el nombre del archivo desde la ruta de origen
    file_name = os.path.basename(origin)
    
    # Crear la ruta completa de destino para el archivo
    dest_file = os.path.join(folder, file_name)

    # Verificar si ya existe un archivo con el mismo nombre en la carpeta destino
    if os.path.exists(dest_file):
        try:
            # Eliminar el archivo existente en la carpeta destino
            os.remove(dest_file)
            
        except FileNotFoundError:
            print(f"Error: No se pudo encontrar el archivo {dest_file}.")
        except PermissionError:
            print(f"Error: Permiso denegado al eliminar el archivo en {dest_file}.")
        except Exception as e:
            print(f"Error inesperado al eliminar el archivo: {e}")

    try:
        # Copiar el archivo desde el origen hasta la carpeta destino
        shutil.copy(os.path.join(origin), folder)
        #print(f"Archivo copiado de {origin} a {folder} exitosamente.")
    except FileNotFoundError:
        print(f"Error: No se pudo encontrar el archivo {origin}.")
    except PermissionError:
        print(f"Error: Permiso denegado al copiar el archivo en {folder}.")
    except Exception as e:
        print(f"Error inesperado al copiar el archivo: {e}")



# retorna todos los archivos excel de un directorio
def list_xlsx_to_folder(path):
    archivos_xlsx = []

    for root, dirs, files in os.walk(path):
        for file in files:
            if file.endswith(".xlsx"):
                archivos_xlsx.append(os.path.join(root, file))

    return archivos_xlsx



    
    
